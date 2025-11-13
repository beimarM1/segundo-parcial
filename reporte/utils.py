"""
Utilidades para generar reportes en diferentes formatos con opción de incluir gráficos.
"""

from openpyxl.chart import BarChart, Reference, LineChart
from venta.models import Venta  # ✅ asegúrate que el path sea correcto
import tempfile
from reportlab.pdfgen import canvas
from io import BytesIO
import os
import matplotlib.pyplot as plt
from django.utils import timezone
from django.db.models import Sum, Count
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ===========================================================
# 📊 GENERADOR DE GRÁFICOS MATPLOTLIB
# ===========================================================
def generar_grafico_ventas_por_mes(datos):
    """
    Genera un gráfico de barras con las ventas agrupadas por mes.
    Devuelve la ruta temporal del gráfico generado.
    """
    if not datos.get("ventas_detalle"):
        return None

    # Contar ventas por mes
    from collections import Counter
    from datetime import datetime

    conteo_por_mes = Counter()

    for v in datos["ventas_detalle"]:
        fecha = datetime.strptime(v["fecha"], "%d/%m/%Y %H:%M")
        clave_mes = fecha.strftime("%b %Y")
        conteo_por_mes[clave_mes] += v["total"]

    meses = list(conteo_por_mes.keys())
    totales = list(conteo_por_mes.values())

    plt.figure(figsize=(8, 4))
    plt.bar(meses, totales, color="#3498db")
    plt.title("Ventas por Mes", fontsize=14, color="#2c3e50")
    plt.xlabel("Mes", fontsize=12)
    plt.ylabel("Total Ventas ($)", fontsize=12)
    plt.xticks(rotation=30, ha="right")
    plt.tight_layout()

    # Guardar imagen temporal
    buffer_img = BytesIO()
    plt.savefig(buffer_img, format="png", dpi=100)
    plt.close()
    buffer_img.seek(0)
    return buffer_img


# ===========================================================
# 🧾 GENERADOR DE PDF
# ===========================================================
def generar_reporte_ventas_pdf(
    datos, fecha_inicio=None, fecha_fin=None, incluir_graficos=True
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#2c3e50"),
        alignment=TA_CENTER,
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=12,
        textColor=colors.HexColor("#7f8c8d"),
        alignment=TA_CENTER,
        spaceAfter=10,
    )

    # === ENCABEZADO ===
    elements.append(Paragraph("SmartSales365", title_style))
    elements.append(Paragraph("Reporte de Ventas", subtitle_style))

    if fecha_inicio and fecha_fin:
        periodo = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periodo, subtitle_style))
    elements.append(Spacer(1, 20))

    # === TABLA DE RESUMEN ===
    resumen_data = [
        ["Métrica", "Valor"],
        ["Total de Ventas", f"${datos['total_ventas']:.2f}"],
        ["Cantidad de Ventas", str(datos["cantidad_ventas"])],
        ["Ticket Promedio", f"${datos['ticket_promedio']:.2f}"],
        ["Productos Vendidos", str(datos["productos_vendidos"])],
    ]
    resumen_table = Table(resumen_data, colWidths=[3 * inch, 2 * inch])
    resumen_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 13),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))

    # === GRÁFICO (si aplica) ===
    if incluir_graficos:
        grafico_buffer = generar_grafico_ventas_por_mes(datos)
        if grafico_buffer:
            elements.append(Paragraph("Gráfico de Ventas por Mes", styles["Heading2"]))
            elements.append(Spacer(1, 10))
            img = Image(grafico_buffer, width=6 * inch, height=3 * inch)
            elements.append(img)
            elements.append(Spacer(1, 20))

    # === DETALLE DE VENTAS ===
    if datos.get("ventas_detalle"):
        elements.append(Paragraph("Detalle de Ventas", styles["Heading2"]))
        elements.append(Spacer(1, 10))

        ventas_data = [["ID", "Cliente", "Fecha", "Total", "Estado"]]
        for venta in datos["ventas_detalle"]:
            ventas_data.append(
                [
                    str(venta["id"]),
                    venta["usuario"],
                    venta["fecha"],
                    f"${venta['total']:.2f}",
                    venta["estado"],
                ]
            )
        ventas_table = Table(
            ventas_data,
            colWidths=[0.7 * inch, 2 * inch, 1.5 * inch, 1.2 * inch, 1.2 * inch],
        )
        ventas_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.lightgrey],
                    ),
                ]
            )
        )
        elements.append(ventas_table)

    # === PIE ===
    elements.append(Spacer(1, 20))
    fecha_gen = timezone.now().strftime("%d/%m/%Y %H:%M:%S")
    pie = Paragraph(
        f"<i>Reporte generado el {fecha_gen}</i>",
        ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_RIGHT,
        ),
    )
    elements.append(pie)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ===========================================================
# 🧾 GENERADOR DE EXCEL
# ===========================================================
def generar_reporte_ventas_excel(
    datos, fecha_inicio=None, fecha_fin=None, incluir_graficos=True
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    header_fill = PatternFill(
        start_color="3498db", end_color="3498db", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="2c3e50")

    ws["A1"] = "SmartSales365 - Reporte de Ventas"
    ws["A1"].font = title_font
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    if fecha_inicio and fecha_fin:
        ws["A2"] = (
            f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        )
        ws.merge_cells("A2:E2")
        ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    ws[f"A{row}"] = "RESUMEN GENERAL"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws.merge_cells(f"A{row}:B{row}")

    row += 1
    resumen_items = [
        ("Total de Ventas", datos["total_ventas"]),
        ("Cantidad de Ventas", datos["cantidad_ventas"]),
        ("Ticket Promedio", datos["ticket_promedio"]),
        ("Productos Vendidos", datos["productos_vendidos"]),
    ]
    for label, valor in resumen_items:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = valor
        ws[f"B{row}"].number_format = "#,##0.00"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    if datos.get("ventas_detalle"):
        row += 2
        ws[f"A{row}"] = "DETALLE DE VENTAS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        headers = ["ID", "Cliente", "Fecha", "Total", "Estado"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row += 1
        for venta in datos["ventas_detalle"]:
            ws.cell(row=row, column=1, value=venta["id"])
            ws.cell(row=row, column=2, value=venta["usuario"])
            ws.cell(row=row, column=3, value=venta["fecha"])
            ws.cell(row=row, column=4, value=venta["total"])
            ws.cell(row=row, column=5, value=venta["estado"])
            row += 1

    # Ajuste de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    # === Insertar gráfico (opcional) ===
    if incluir_graficos:
        grafico_buffer = generar_grafico_ventas_por_mes(datos)
        if grafico_buffer:
            img = XLImage(grafico_buffer)
            img.anchor = f"G5"
            ws.add_image(img)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ===========================================================
# 📈 OBTENCIÓN DE DATOS
# ===========================================================
def generar_datos_reporte_ventas(fecha_inicio=None, fecha_fin=None):
    """
    Obtiene datos agregados de ventas, filtrados por fecha.
    """
    from venta.models import Venta, DetalleVenta
    from django.db.models import Q

    ventas_query = Venta.objects.filter(estado="pagado")
    if fecha_inicio:
        ventas_query = ventas_query.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        ventas_query = ventas_query.filter(fecha__lte=fecha_fin)

    total_ventas = ventas_query.aggregate(total=Sum("total"))["total"] or 0
    cantidad_ventas = ventas_query.count()
    ticket_promedio = total_ventas / cantidad_ventas if cantidad_ventas > 0 else 0

    productos_vendidos = (
        DetalleVenta.objects.filter(venta__in=ventas_query).aggregate(
            total=Sum("cantidad")
        )["total"]
        or 0
    )

    ventas_detalle = []
    for venta in ventas_query.order_by("-fecha")[:50]:
        ventas_detalle.append(
            {
                "id": venta.id,
                "usuario": venta.usuario.username if venta.usuario else "N/A",
                "fecha": venta.fecha.strftime("%d/%m/%Y %H:%M"),
                "total": float(venta.total),
                "estado": venta.get_estado_display(),
            }
        )

    return {
        "total_ventas": float(total_ventas),
        "cantidad_ventas": cantidad_ventas,
        "ticket_promedio": float(ticket_promedio),
        "productos_vendidos": productos_vendidos,
        "ventas_detalle": ventas_detalle,
    }


# ===========================================================
# 📈 Productos
# ===========================================================


def generar_reporte_productos_pdf(datos):
    """
    Genera un reporte de productos en formato PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("SmartSales365 - Reporte de Productos", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 20))

    # Resumen general
    resumen = [
        ["Total de productos", str(datos["total_productos"])],
        ["Valor total del inventario", f"${datos['valor_inventario']:.2f}"],
    ]
    resumen_table = Table(resumen, colWidths=[250, 250])
    resumen_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))

    # Detalle de productos
    products_data = [["ID", "Nombre", "Marca", "Categoría", "Precio", "Stock"]]
    for p in datos["productos"]:
        products_data.append(
            [
                p["id"],
                p["nombre"],
                p["marca"],
                p["categoria"],
                f"${p['precio']:.2f}",
                p["stock"],
            ]
        )

    table = Table(products_data, colWidths=[40, 120, 80, 80, 80, 60])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer



#####################################################################
#productos excel
############################################################

def generar_reporte_productos_excel(datos_reporte):
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezado
    headers = ["ID", "Nombre", "Marca", "Categoría", "Precio", "Stock", "Estado"]
    ws.append(headers)

    # Filas de datos
    for p in datos_reporte.get("productos", []):
        ws.append([
            p["id"],
            p["nombre"],
            p["marca"],
            p["categoria"],
            p["precio"],
            p["stock"],
            p["estado"],
        ])

    # Ajustar ancho de columnas
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    wb.save(buffer)
    buffer.seek(0)
    return buffer



#############################################################################
#      clientes pdf
#################################################################################


def generar_reporte_clientes_pdf(datos_reporte):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, height - 50, "Reporte de Clientes")

    p.setFont("Helvetica", 10)
    y = height - 100

    clientes = datos_reporte.get("clientes", [])
    p.drawString(50, y, f"Total de clientes: {datos_reporte.get('total_clientes', 0)}")
    y -= 30

    headers = ["ID", "Username", "Email", "Compras", "Total Compras", "Registro"]
    for i, h in enumerate(headers):
        p.drawString(50 + i * 90, y, h)
    y -= 20

    for cliente in clientes:
        if y < 100:  # nueva página si se llena
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 100

        p.drawString(50, y, str(cliente["id"]))
        p.drawString(140, y, cliente["username"])
        p.drawString(240, y, cliente["email"][:20])
        p.drawString(360, y, str(cliente["cantidad_compras"]))
        p.drawString(430, y, f"${cliente['total_compras']:.2f}")
        p.drawString(500, y, cliente["fecha_registro"])
        y -= 18

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer


############################################################################################
#    cliente excel
#############################################################################################

def generar_reporte_clientes_excel(datos_reporte, incluir_graficos=True):
    """
    Genera un archivo Excel con los datos de clientes.
    Si incluir_graficos=True, añade un gráfico de barras de total de compras.
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # ==============================
    # ENCABEZADOS
    # ==============================
    headers = [
        "ID",
        "Username",
        "Email",
        "Cantidad Compras",
        "Total Compras (Bs.)",
        "Fecha Registro",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ==============================
    # FILAS DE DATOS
    # ==============================
    clientes = datos_reporte.get("clientes", [])
    for c in clientes:
        ws.append([
            c["id"],
            c["username"],
            c["email"],
            c["cantidad_compras"],
            c["total_compras"],
            c["fecha_registro"],
        ])

    # ==============================
    # AJUSTAR ANCHO DE COLUMNAS
    # ==============================
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    # ==============================
    # RESUMEN
    # ==============================
    total_clientes = datos_reporte.get("total_clientes", 0)
    ws.append([])
    ws.append(["", "", "", "TOTAL CLIENTES:", total_clientes])
    last_row = ws.max_row
    ws.cell(row=last_row, column=4).font = Font(bold=True)
    ws.cell(row=last_row, column=5).font = Font(bold=True)

    # ==============================
    # GRÁFICO DE BARRAS (opcional)
    # ==============================
    if incluir_graficos and clientes:
        chart = BarChart()
        chart.title = "Total de compras por cliente"
        chart.x_axis.title = "Clientes"
        chart.y_axis.title = "Monto total (Bs.)"

        # Determinar rango de datos
        start_row = 2
        end_row = 1 + len(clientes)
        data = Reference(ws, min_col=5, min_row=1, max_row=end_row)  # Total Compras
        cats = Reference(ws, min_col=2, min_row=2, max_row=end_row)  # Username

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.style = 10  # estilo azul
        chart.height = 8
        chart.width = 18

        # Ubicar el gráfico a partir de la columna H (8)
        ws.add_chart(chart, f"H4")

    # Guardar en buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer


#############################################################################################
#       pdf inventario
#############################################################################################3


def generar_reporte_inventario_pdf(datos_reporte):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 14)
    p.drawString(180, height - 50, "Reporte de Inventario")

    p.setFont("Helvetica", 10)
    y = height - 100

    # Resumen general
    p.drawString(50, y, f"Total de productos: {datos_reporte.get('total_productos', 0)}")
    y -= 20
    p.drawString(50, y, f"Productos bajo stock (<10): {datos_reporte.get('productos_bajo_stock', 0)}")
    y -= 20
    p.drawString(50, y, f"Productos sin stock: {datos_reporte.get('productos_sin_stock', 0)}")
    y -= 20
    p.drawString(50, y, f"Valor total del inventario: Bs. {datos_reporte.get('valor_total_inventario', 0):,.2f}")
    y -= 40

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Productos con bajo stock:")
    y -= 20
    p.setFont("Helvetica", 10)

    # Listado de productos con bajo stock
    for pbs in datos_reporte.get("productos_bajo_stock_detalle", []):
        if y < 80:  # salto de página si se llena
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 100
        p.drawString(60, y, f"- {pbs['nombre']} | Stock: {pbs['stock']} | Precio: Bs. {pbs['precio']:.2f}")
        y -= 15

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer




#########################################################################################
#    inventario excel
##################################################################################
def generar_reporte_inventario_excel(datos_reporte, incluir_graficos=True):
    """
    Genera un archivo Excel para el reporte de inventario.
    Si incluir_graficos=True, añade un gráfico de barras de productos con bajo stock.
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # ===========================
    # ENCABEZADOS
    # ===========================
    headers = ["Nombre", "Stock", "Precio (Bs.)"]
    ws.append(headers)
    header_font = Font(bold=True)

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # ===========================
    # FILAS DE PRODUCTOS BAJO STOCK
    # ===========================
    productos_bajo = datos_reporte.get("productos_bajo_stock_detalle", [])
    for p in productos_bajo:
        ws.append([p["nombre"], p["stock"], p["precio"]])

    # ===========================
    # RESUMEN GENERAL
    # ===========================
    ws.append([])
    ws.append(["", ""])
    resumen_start = ws.max_row + 1

    resumen_data = [
        ["Total productos", datos_reporte.get("total_productos", 0)],
        ["Productos bajo stock (<10)", datos_reporte.get("productos_bajo_stock", 0)],
        ["Productos sin stock", datos_reporte.get("productos_sin_stock", 0)],
        ["Valor total inventario (Bs.)", datos_reporte.get("valor_total_inventario", 0)],
    ]

    for row in resumen_data:
        ws.append(row)

    for i in range(1, 3):
        ws.column_dimensions[get_column_letter(i)].width = 30

    # ===========================
    # GRÁFICO DE STOCK (opcional)
    # ===========================
    if incluir_graficos and productos_bajo:
        chart = BarChart()
        chart.title = "Productos con Bajo Stock"
        chart.x_axis.title = "Producto"
        chart.y_axis.title = "Unidades en stock"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(productos_bajo) + 1)  # Stock
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(productos_bajo) + 1)  # Nombres

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.style = 11
        chart.height = 8
        chart.width = 18

        # Colocar el gráfico más abajo del resumen
        ws.add_chart(chart, f"E4")

    wb.save(buffer)
    buffer.seek(0)
    return buffer



################################################################################################33
#  pdf financiero
###########################################################################################################
def generar_reporte_financiero_pdf(datos_reporte, incluir_graficos=True):
    """
    Genera un archivo PDF para el reporte financiero.
    Si incluir_graficos=True, añade un gráfico de ingresos.
    """
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # ==============================
    # ENCABEZADO
    # ==============================
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2, height - 50, "Reporte Financiero")

    p.setFont("Helvetica", 10)
    y = height - 100

    # ==============================
    # DATOS GENERALES
    # ==============================
    ingresos = datos_reporte.get("ingresos_totales", 0)
    transacciones = datos_reporte.get("cantidad_transacciones", 0)
    ticket = datos_reporte.get("ticket_promedio", 0)
    periodo = datos_reporte.get("periodo", {})

    p.drawString(50, y, f"Ingresos Totales: Bs. {ingresos:,.2f}")
    y -= 20
    p.drawString(50, y, f"Cantidad de Transacciones: {transacciones}")
    y -= 20
    p.drawString(50, y, f"Ticket Promedio: Bs. {ticket:,.2f}")
    y -= 20
    p.drawString(50, y, f"Periodo: {periodo.get('fecha_inicio', 'N/A')} → {periodo.get('fecha_fin', 'N/A')}")
    y -= 40

    # ==============================
    # GRÁFICO (opcional)
    # ==============================
    if incluir_graficos and ingresos > 0:
        try:
            # Creamos un gráfico simple con Matplotlib
            fig, ax = plt.subplots(figsize=(4, 2.5))
            ax.bar(
                ["Ingresos Totales", "Ticket Promedio"],
                [ingresos, ticket],
                color=["#2a3964", "#880000"],
            )
            ax.set_title("Resumen Financiero")
            ax.set_ylabel("Monto (Bs.)")
            ax.grid(axis="y", linestyle="--", alpha=0.7)

            # Guardar imagen temporal
            tmpfile = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            plt.savefig(tmpfile.name, bbox_inches="tight")
            plt.close(fig)

            # Insertar imagen en PDF
            p.drawImage(tmpfile.name, 70, y - 200, width=5.8 * inch, height=2.6 * inch)
            y -= 220
        except Exception as e:
            print("⚠️ No se pudo generar el gráfico: - utils.py:783", e)

    # ==============================
    # PIE DE PÁGINA
    # ==============================
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, 40, "SmartSales365 © Reporte financiero generado automáticamente")

    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer




#############################################################
#    finaciero excel
#################################################################


def generar_reporte_financiero_excel(datos_reporte, incluir_graficos=True):
    """
    Genera un reporte financiero con resumen y gráfico de tendencia mensual de ingresos.
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financiero"

    # =============================
    # ENCABEZADO PRINCIPAL
    # =============================
    ws["A1"] = "Reporte Financiero - SmartSales365"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # =============================
    # RESUMEN GLOBAL
    # =============================
    ws.append([])
    ws.append(["Indicador", "Valor (Bs.)"])
    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    ingresos = float(datos_reporte.get("ingresos_totales", 0))
    transacciones = int(datos_reporte.get("cantidad_transacciones", 0))
    ticket = float(datos_reporte.get("ticket_promedio", 0))
    periodo = datos_reporte.get("periodo", {})

    ws.append(["Ingresos Totales", ingresos])
    ws.append(["Cantidad de Transacciones", transacciones])
    ws.append(["Ticket Promedio", ticket])
    ws.append(["Periodo Inicio", periodo.get("fecha_inicio", "N/A")])
    ws.append(["Periodo Fin", periodo.get("fecha_fin", "N/A")])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    for row in ws.iter_rows(min_row=3, max_row=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # =============================
    # TENDENCIA MENSUAL DE INGRESOS
    # =============================
    ws2 = wb.create_sheet("Tendencia Ingresos")

    ws2["A1"] = "Mes"
    ws2["B1"] = "Ingresos (Bs.)"
    ws2["A1"].font = ws2["B1"].font = Font(bold=True)
    ws2["A1"].alignment = ws2["B1"].alignment = Alignment(horizontal="center")

    # Consultar datos mensuales
    ventas_mensuales = (
        Venta.objects.filter(estado="pagado")
        .values_list("fecha__year", "fecha__month")
        .annotate(total_mes=Sum("total"))
        .order_by("fecha__year", "fecha__month")
    )

    if not ventas_mensuales:
        ws2.append(["Sin datos de ventas", 0])
    else:
        for year, month, total in ventas_mensuales:
            nombre_mes = f"{month:02d}/{year}"
            ws2.append([nombre_mes, float(total)])

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 20

    # =============================
    # GRÁFICO DE BARRAS / LÍNEA
    # =============================
    if incluir_graficos and ventas_mensuales:
        chart = LineChart()
        chart.title = "Evolución Mensual de Ingresos"
        chart.x_axis.title = "Mes"
        chart.y_axis.title = "Monto (Bs.)"

        data = Reference(ws2, min_col=2, min_row=1, max_row=len(ventas_mensuales) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(ventas_mensuales) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 12
        chart.height = 8
        chart.width = 18

        ws2.add_chart(chart, "D3")

    # =============================
    # PIE DE PÁGINA
    # =============================
    ws["A10"] = "SmartSales365 © Reporte financiero generado automáticamente"
    ws["A10"].font = Font(italic=True, size=9, color="888888")

    wb.save(buffer)
    buffer.seek(0)
    return buffer
